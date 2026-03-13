import openpyxl
wb = openpyxl.load_workbook('1-3.成绩单-门店维度2026-03-01-2026-03-01.xlsx')
ws = wb.active

# Row 2 = sub-headers, Row 3+ = data
# Cols 1-7: 门店id, 所属集团, 所属大区, 所属小区, 门店全称, 自营/合营, 总店
# Cols 8-24: 合计 (17 cols)
# Cols 25-42: 收购业务 (18 cols)
# Cols 43-57: 委拍业务 (15 cols)

rows = []
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    vals = list(row)
    if not vals[0] and not vals[4]:  # skip empty rows
        continue
    
    def fmt(v):
        if v is None or v == '' or v == 'None':
            return "''"
        v = str(v)
        # Escape single quotes
        v = v.replace("'", "\\'")
        return f"'{v}'"
    
    # Build object: s0=门店id, s1=集团, s2=大区, s3=小区, s4=门店全称, s5=自营/合营, s6=总店
    # Then t0-t16 for 合计(17), sg0-sg17 for 收购(18), wp0-wp14 for 委拍(15)
    parts = []
    # Info columns
    info_keys = ['s0','s1','s2','s3','s4','s5','s6']
    for i, k in enumerate(info_keys):
        parts.append(f"{k}:{fmt(vals[i])}")
    
    # 合计 cols 8-24 (index 7-23), 17 cols
    for i in range(17):
        parts.append(f"t{i}:{fmt(vals[7+i])}")
    
    # 收购 cols 25-42 (index 24-41), 18 cols
    for i in range(18):
        parts.append(f"sg{i}:{fmt(vals[24+i])}")
    
    # 委拍 cols 43-57 (index 42-56), 15 cols
    for i in range(15):
        parts.append(f"wp{i}:{fmt(vals[42+i])}")
    
    rows.append('{' + ','.join(parts) + '}')

# Generate columns
# 合计 sub-headers (row 2, cols 8-24)
total_labels = []
for c in range(8, 25):
    total_labels.append(ws.cell(row=2, column=c).value or '')

purchase_labels = []
for c in range(25, 43):
    purchase_labels.append(ws.cell(row=2, column=c).value or '')

consign_labels = []
for c in range(43, 58):
    consign_labels.append(ws.cell(row=2, column=c).value or '')

# Build storeCols
cols = []
# Info cols
cols.append("{k:'s0',l:'门店id'}")
cols.append("{k:'s1',l:'所属集团'}")
cols.append("{k:'s2',l:'所属大区'}")
cols.append("{k:'s3',l:'所属小区'}")
cols.append("{k:'s4',l:'门店全称'}")
cols.append("{k:'s5',l:'自营/合营'}")
cols.append("{k:'s6',l:'总店'}")

# 合计
for i, lbl in enumerate(total_labels):
    extra = ""
    if '汇和毛利' in lbl:
        extra = ",cls:'changed'"
    elif '总毛利' in lbl and '台均' not in lbl:
        extra = ",cls:'changed'"
    cols.append("{k:'t" + str(i) + "',l:'" + lbl + "'" + extra + "}")

# 收购
for i, lbl in enumerate(purchase_labels):
    extra = ""
    if '汇和毛利' in lbl:
        extra = ",cls:'changed'"
    elif '总毛利' in lbl and '台均' not in lbl:
        extra = ",cls:'changed'"
    cols.append("{k:'sg" + str(i) + "',l:'" + lbl + "'" + extra + "}")

# 委拍
for i, lbl in enumerate(consign_labels):
    extra = ""
    if '汇和毛利' in lbl:
        extra = ",cls:'changed'"
    elif '总毛利' in lbl and '台均' not in lbl:
        extra = ",cls:'changed'"
    cols.append("{k:'wp" + str(i) + "',l:'" + lbl + "'" + extra + "}")

print("// Store cols")
print("const storeCols=[" + ",".join(cols) + "];")
print()
print("// Store groups")
print("const storeGroups=[{name:'合计',span:17},{name:'收购业务',span:18},{name:'委拍业务',span:15}];")
print()
print("// Store data")
print("const storeData=[")
for r in rows:
    print(r + ",")
print("];")
